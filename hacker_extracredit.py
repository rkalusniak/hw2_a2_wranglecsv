"""
Hacker Extra Credit

Created: June 12 19:21:33 2023

by: Rachel Kalusniak
"""
#Import libraries
import pandas as pd
import numpy as np
import csv
import openpyxl
from openpyxl.styles import NamedStyle
from pathlib import Path



#Create the path to the data and output
data_dir = Path('data', 'raw', 'morelogs')
output_dir = Path('output', 'extracredit')



#Create Excel number formats. It doesn't like doing this multiple times in loop below
#Create temp format
temp_style = NamedStyle(name='temp_style', number_format='0.0')

#Create date format
date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD HH:MM:SS')



#Create list of unique streams
#Initialize list
stream = []

#Get all the files in folder. Record file name and stem in list
for f in data_dir.rglob('*.csv'):
    stream.append(f.stem.split('-')[0])

#Create list of unique stream names
stream_unique = np.unique(stream)

#Loop over each unique stream
for s in stream_unique:


    #Create and save the workbook
    wb = openpyxl.Workbook()
    xl_file = Path(output_dir, s + '.xlsx').as_posix()
    wb.save(xl_file)


    # Loop over files in folder
    for f in data_dir.glob(s + '*.csv'):
        # Create a string with path from project folder and use as.posix to flip to forward slash.
        strfile = Path(data_dir, f.name).as_posix()

        # Read in csv file to dataframe
        csv_df = pd.read_csv(strfile, header=None, names=['datetime', 'scale', 'temperature'],
                             parse_dates=['datetime'])

        # Write each dataframe to a sheet
        with pd.ExcelWriter(xl_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            csv_df.to_excel(writer, sheet_name=f.stem, index=False)


        # Write summary statistics to excel
        # Reopen workbook and set worksheet
        wb = openpyxl.load_workbook(xl_file)
        ws = wb[f.stem]

        # Find data length
        ws_length = str(len(csv_df.index) + 1)

        # Min temp
        ws['G2'] = 'min_temp'
        ws['H2'] = '=MIN(C2:C' + ws_length + ')'
        ws['H2'].style = temp_style

        # Max temp
        ws['G3'] = 'max_temp'
        ws['H3'] = '=MAX(C2:C' + ws_length + ')'
        ws['H3'].style = temp_style

        # Mean temp
        ws['G4'] = 'mean_temp'
        ws['H4'] = '=AVERAGE(C2:C' + ws_length + ')'
        ws['H4'].style = temp_style

        # Min date
        ws['G6'] = 'min_date'
        ws['H6'] = '=MIN(A2:A' + ws_length + ')'
        ws['H6'].style = date_style

        # Max date
        ws['G7'] = 'max_date'
        ws['H7'] = '=MAX(A2:A' + ws_length + ')'
        ws['H7'].style = date_style

        # Save workbook with changes
        wb.save(xl_file)



#Print that the process is done
print('The data was was written to the excel sheet successfully.')